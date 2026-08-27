from __future__ import annotations

import argparse
import csv
import json
import math
import sys
import threading
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Iterable
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import Download as Main
from Config.MarketCalendar import (
    HolidayName,
    IsMarketSession,
    IsTradingDate,
    MarketClosedReason,
    ShouldRunSingleOffmarketCheck,
    ShouldStartLiveTick,
)
from Strategies.G01 import Core as base
from Strategies.G01 import Gold as gold
from LiveTick.session import RuntimeSession, is_pid_running, load_session, request_stop


IST = ZoneInfo("Asia/Kolkata")
MARKET_OPEN = "09:15"
MARKET_CLOSE = "15:30"
EOD_EXIT_TIME = "15:25"
DEFAULT_SYMBOLS = ("CGPOWER", "HDFCBANK")
DEFAULT_DATA_FOLDER = ROOT / "Data"
DEFAULT_REPORT_DIR = ROOT / "Paper" / "Reports"
DEFAULT_TICK_ROOT = ROOT / "TickData"
ANSI_GREEN = "\033[92m"
ANSI_RESET = "\033[0m"


@dataclass(frozen=True)
class PaperConfig:
    initial_balance: float = 1000.0
    leverage: float = 5.0
    max_open_positions: int = 1
    entry_grace_minutes: int = 2
    max_data_staleness_minutes: int = 3
    brokerage_rate: float = 0.0003
    brokerage_cap_per_order: float = 20.0
    stt_sell_side_rate: float = 0.00025
    exchange_txn_rate: float = 0.0000307
    sebi_turnover_rate: float = 0.000001
    stamp_buy_side_rate: float = 0.00003
    gst_rate: float = 0.18
    slippage_bps: float = 0.0


def now_ist() -> datetime:
    return datetime.now(IST).replace(tzinfo=None)


def parse_time(value: str):
    return datetime.strptime(value, "%H:%M").time()


def clean_symbol(symbol: str) -> str:
    value = str(symbol).strip().upper()
    if ":" in value:
        value = value.split(":", 1)[1]
    return value.replace("-EQ", "")


def is_market_session(ts: datetime) -> bool:
    return IsMarketSession(ts)


def is_trading_day(ts: datetime) -> bool:
    return IsTradingDate(ts)


def should_start_live_tick(ts: datetime, manage_live_tick: bool = True) -> bool:
    return ShouldStartLiveTick(ts, manage_live_tick=manage_live_tick)


def should_run_single_offmarket_check(ts: datetime) -> bool:
    return ShouldRunSingleOffmarketCheck(ts)


def offmarket_live_tick_message(ts: datetime) -> str:
    reason = MarketClosedReason(ts) or "LiveTick auto-start is disabled"
    holiday = HolidayName(ts)
    if holiday:
        return f"{reason}; LiveTick will not start. Running one local paper check only."
    if not is_trading_day(ts):
        return f"{reason}; LiveTick will not start. Running one local paper check only."
    if ts.time() > parse_time(MARKET_CLOSE):
        return f"{reason}; LiveTick will not start. Running one local paper check only."
    return "LiveTick auto-start is disabled; paper trader will only use existing local CSV candles."


def latest_expected_closed_minute(ts: datetime) -> datetime | None:
    if not is_market_session(ts):
        return None
    candidate = ts.replace(second=0, microsecond=0) - timedelta(minutes=1)
    market_open = ts.replace(hour=9, minute=15, second=0, microsecond=0)
    if candidate < market_open:
        return None
    return candidate


def report_dir(path: str | Path = DEFAULT_REPORT_DIR) -> Path:
    directory = Path(path)
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def state_path(report_folder: str | Path = DEFAULT_REPORT_DIR) -> Path:
    return report_dir(report_folder) / "gold_paper_state.json"


def events_path(report_folder: str | Path = DEFAULT_REPORT_DIR) -> Path:
    return report_dir(report_folder) / "gold_paper_events.csv"


def trades_path(report_folder: str | Path = DEFAULT_REPORT_DIR) -> Path:
    return report_dir(report_folder) / "gold_paper_trades.csv"


def report_xlsx_path(report_folder: str | Path = DEFAULT_REPORT_DIR, current: datetime | None = None) -> Path:
    day = (current or now_ist()).strftime("%Y-%m-%d")
    return report_dir(report_folder) / f"GoldPaperTrades_{day}.xlsx"


def fallback_report_xlsx_path(primary_path: str | Path) -> Path:
    return fallback_report_file_path(primary_path)


def fallback_report_file_path(primary_path: str | Path) -> Path:
    primary = Path(primary_path)
    return primary.with_name(f"{primary.stem}_live{primary.suffix}")


def timeframe_path(symbol: str, timeframe: str, data_folder: str | Path = DEFAULT_DATA_FOLDER) -> Path:
    symbol = clean_symbol(symbol)
    return Path(data_folder) / symbol / f"{symbol}_{timeframe}.csv"


def load_candles(symbol: str, timeframe: str, data_folder: str | Path = DEFAULT_DATA_FOLDER) -> pd.DataFrame:
    path = timeframe_path(symbol, timeframe, data_folder)
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=Main.FINAL_COLUMNS)
    return Main.normalize_candles(pd.read_csv(path, parse_dates=["Datetime"]))


def load_state(
    symbols: Iterable[str],
    config: PaperConfig,
    report_folder: str | Path = DEFAULT_REPORT_DIR,
    *,
    reset: bool = False,
) -> dict[str, Any]:
    path = state_path(report_folder)
    today = now_ist().date().isoformat()
    cleaned_symbols = [clean_symbol(symbol) for symbol in symbols]
    if reset or not path.exists():
        state = {
            "session_date": today,
            "started_at": now_ist().isoformat(timespec="seconds"),
            "symbols": cleaned_symbols,
            "config": asdict(config),
            "balance": float(config.initial_balance),
            "open_positions": {},
            "seen_signal_keys": [],
            "missed_signal_keys": [],
            "last_status": "",
        }
        save_state(state, report_folder)
        return state

    state = json.loads(path.read_text(encoding="utf-8"))
    if state.get("session_date") != today:
        state.update(
            {
                "session_date": today,
                "started_at": now_ist().isoformat(timespec="seconds"),
                "symbols": cleaned_symbols,
                "open_positions": {},
                "seen_signal_keys": [],
                "missed_signal_keys": [],
                "last_status": "",
            }
        )
    state.setdefault("config", asdict(config))
    state.setdefault("balance", float(config.initial_balance))
    state.setdefault("open_positions", {})
    state.setdefault("seen_signal_keys", [])
    state.setdefault("missed_signal_keys", [])
    state["symbols"] = cleaned_symbols
    save_state(state, report_folder)
    return state


def save_state(state: dict[str, Any], report_folder: str | Path = DEFAULT_REPORT_DIR) -> None:
    path = state_path(report_folder)
    path.write_text(json.dumps(state, indent=2, default=str), encoding="utf-8")


def append_csv(path: Path, row: dict[str, Any]) -> None:
    try:
        _append_csv_to_path(path, row)
    except PermissionError:
        fallback = fallback_report_file_path(path)
        _append_csv_to_path(fallback, row)
        print(f"Report CSV is locked, wrote live CSV copy instead: {fallback}", flush=True)


def _append_csv_to_path(path: Path, row: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    fieldnames = list(row.keys())
    if path.exists() and path.stat().st_size > 0:
        with path.open("r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            fieldnames = list(dict.fromkeys(list(reader.fieldnames or []) + list(row.keys())))
            rows = list(reader)
    rows.append({key: row.get(key, "") for key in fieldnames})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def append_event(
    event: str,
    symbol: str,
    message: str,
    report_folder: str | Path = DEFAULT_REPORT_DIR,
    extra: dict[str, Any] | None = None,
) -> None:
    row = {
        "timestamp": now_ist().isoformat(timespec="seconds"),
        "event": event,
        "symbol": clean_symbol(symbol),
        "message": message,
    }
    if extra:
        for key, value in extra.items():
            if isinstance(value, (dict, list, tuple, set)):
                row[key] = json.dumps(value, default=str)
            else:
                row[key] = value
    append_csv(events_path(report_folder), row)


def prepare_live_features(df_5m: pd.DataFrame) -> pd.DataFrame:
    df = Main.normalize_candles(df_5m)
    if df.empty:
        return df
    df = df.sort_values("Datetime").reset_index(drop=True)
    df["date"] = df["Datetime"].dt.date.astype(str)
    df["time"] = df["Datetime"].dt.strftime("%H:%M")
    regular = df[(df["time"] >= base.MARKET_OPEN) & (df["time"] <= base.MARKET_CLOSE)].copy()
    if regular.empty:
        return regular
    regular["bar_no"] = regular.groupby("date").cumcount()

    typical = (regular["High"] + regular["Low"] + regular["Close"]) / 3
    volume_cumsum = regular["Volume"].groupby(regular["date"]).cumsum().replace(0, np.nan)
    regular["vwap"] = (typical * regular["Volume"]).groupby(regular["date"]).cumsum() / volume_cumsum
    regular["vwap"] = regular["vwap"].ffill()
    for span in (13, 21, 34, 55):
        regular[f"ema{span}"] = base.ema(regular["Close"], span)
    regular["rsi14"] = base.rsi(regular["Close"], 14)
    regular["atr14"] = base.atr(regular, 14)
    regular["adx_for_signal"] = regular["ADX"] if "ADX" in regular.columns else np.nan
    regular["vol_avg20_samebar"] = regular.groupby("bar_no")["Volume"].transform(
        lambda s: s.shift(1).rolling(20, min_periods=5).mean()
    )
    regular["vol_ratio20"] = regular["Volume"] / regular["vol_avg20_samebar"].replace(0, np.nan)
    regular["prev_close"] = regular["Close"].shift(1)
    return regular.reset_index(drop=True)


def generate_live_signals(df: pd.DataFrame, config: base.StrategyConfig = gold.GOLD_CONFIG) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    long_mask = (
        (df["bar_no"] >= config.long_first_bar)
        & (df["bar_no"] < config.long_last_signal_bar_exclusive)
        & (df["Close"] > df["vwap"])
        & (df["ema21"] > df["ema55"])
        & (df["adx_for_signal"] >= config.adx_min)
        & (df["vol_ratio20"] >= config.volume_ratio_min)
        & (df["rsi14"].between(config.long_rsi_min, config.long_rsi_max))
        & ((df["Low"] <= df["ema21"]) | (df["Low"] <= df["vwap"]))
        & (df["Close"] > df["ema21"])
        & (df["Close"] > df["prev_close"])
    )
    short_mask = (
        (df["bar_no"] >= config.short_first_bar)
        & (df["bar_no"] < config.short_last_signal_bar_exclusive)
        & (df["Close"] < df["vwap"])
        & (df["ema13"] < df["ema34"])
        & (df["adx_for_signal"] >= config.adx_min)
        & (df["vol_ratio20"] >= config.volume_ratio_min)
        & (df["rsi14"].between(config.short_rsi_min, config.short_rsi_max))
        & ((df["High"] >= df["ema13"]) | (df["High"] >= df["vwap"]))
        & (df["Close"] < df["ema13"])
        & (df["Close"] < df["prev_close"])
    )

    signal_rows = pd.concat(
        [
            df.loc[long_mask, ["Datetime", "date", "bar_no"]].assign(direction=1),
            df.loc[short_mask, ["Datetime", "date", "bar_no"]].assign(direction=-1),
        ],
        axis=0,
    ).sort_values(["date", "Datetime", "direction"], ascending=[True, True, False])
    first = signal_rows.groupby("date", as_index=False).first()
    if first.empty:
        return first

    index_lookup = df.reset_index().set_index(["date", "Datetime"])["index"]
    signal_indexes = [int(index_lookup.loc[(row.date, row.Datetime)]) for row in first.itertuples(index=False)]
    first["signal_index"] = signal_indexes
    first["entry_index"] = first["signal_index"] + 1

    entry_times = []
    entries = []
    for row in first.itertuples(index=False):
        signal_dt = pd.Timestamp(row.Datetime).to_pydatetime()
        entry_idx = int(row.entry_index)
        if entry_idx < len(df):
            candidate_dt = pd.Timestamp(df.at[entry_idx, "Datetime"]).to_pydatetime()
            if candidate_dt.date() == signal_dt.date():
                entry_times.append(candidate_dt)
                entries.append(float(df.at[entry_idx, "Open"]))
                continue
        entry_times.append(signal_dt + timedelta(minutes=5))
        entries.append(np.nan)

    first["entry_time"] = entry_times
    first["entry"] = entries
    first["atr14"] = df.loc[first["signal_index"], "atr14"].to_numpy(float)
    first["stop_distance"] = config.stop_atr_multiple * first["atr14"]
    first["stop"] = first["entry"] - first["direction"] * first["stop_distance"]
    first["target"] = first["entry"] + first["direction"] * config.target_r * first["stop_distance"]
    return first.reset_index(drop=True)


def candidate_signal_for_today(df_5m: pd.DataFrame, current: datetime) -> pd.Series | None:
    live = prepare_live_features(df_5m)
    if live.empty or live["date"].nunique() < 25:
        return None

    regime = gold.daily_regime_table(live)
    tradeable_dates = set(regime.loc[regime["regime_tradeable"], "date"])
    today = current.date().isoformat()
    if today not in tradeable_dates:
        return None

    signals = generate_live_signals(live, gold.GOLD_CONFIG)
    signals = signals[signals["date"].isin(tradeable_dates)].copy()
    if signals.empty:
        return None
    strength = gold.signal_strength_table(live, signals, gold.GOLD_CONFIG)
    final = strength[
        (strength["signal_strength"] >= gold.MIN_SIGNAL_STRENGTH)
        & (strength["strength_trigger_component"] >= gold.MIN_TRIGGER_COMPONENT)
    ].copy()
    final = final[final["date"] == today].sort_values("Datetime")
    if final.empty:
        return None
    return final.iloc[0]


def latest_entry_candle(df_1m: pd.DataFrame, entry_time: datetime, grace_minutes: int) -> pd.Series | None:
    if df_1m.empty:
        return None
    rows = df_1m[
        (df_1m["Datetime"] >= entry_time)
        & (df_1m["Datetime"] <= entry_time + timedelta(minutes=max(0, grace_minutes)))
    ].copy()
    if rows.empty:
        return None
    return rows.iloc[0]


def signal_key(symbol: str, signal: pd.Series) -> str:
    return f"{clean_symbol(symbol)}|{pd.Timestamp(signal['Datetime']).isoformat()}|{int(signal['direction'])}"


def signal_alert_message(symbol: str, signal: pd.Series, prefix: str = "SIGNAL HIT") -> str:
    direction = int(signal["direction"])
    side = "Long" if direction == 1 else "Short"
    signal_time = pd.Timestamp(signal["Datetime"]).strftime("%H:%M")
    entry_time = pd.Timestamp(signal["entry_time"]).strftime("%H:%M")
    strength = float(signal.get("signal_strength", 0.0))
    band = str(signal.get("strength_band", ""))
    return (
        f"{clean_symbol(symbol)}: {prefix} {side} signal={signal_time} entry={entry_time} "
        f"strength={strength:.1f} band={band}"
    )


def apply_slippage(price: float, direction: int, side: str, config: PaperConfig) -> float:
    if config.slippage_bps <= 0:
        return float(price)
    sign = 1 if (side == "entry" and direction == 1) or (side == "exit" and direction == -1) else -1
    return float(price) * (1 + sign * config.slippage_bps / 10000)


def open_position(
    symbol: str,
    signal: pd.Series,
    entry_candle: pd.Series,
    state: dict[str, Any],
    config: PaperConfig,
) -> tuple[dict[str, Any] | None, str]:
    direction = int(signal["direction"])
    raw_entry = float(entry_candle["Open"])
    entry = apply_slippage(raw_entry, direction, "entry", config)
    balance = float(state.get("balance", config.initial_balance))
    qty = int((balance * config.leverage) // entry)
    if qty <= 0:
        return None, f"Insufficient paper balance for whole-share {symbol} entry at {entry:.2f}."

    stop_distance = float(signal["stop_distance"])
    stop = entry - direction * stop_distance
    target = entry + direction * gold.GOLD_CONFIG.target_r * stop_distance
    position = {
        "symbol": clean_symbol(symbol),
        "date": str(signal["date"]),
        "signal_key": signal_key(symbol, signal),
        "direction": direction,
        "side": "Long" if direction == 1 else "Short",
        "signal_time": pd.Timestamp(signal["Datetime"]).isoformat(),
        "entry_time": pd.Timestamp(entry_candle["Datetime"]).isoformat(),
        "planned_entry_time": pd.Timestamp(signal["entry_time"]).isoformat(),
        "entry": round(entry, 4),
        "raw_entry": round(raw_entry, 4),
        "qty": qty,
        "stop": round(stop, 4),
        "target": round(target, 4),
        "stop_distance": round(stop_distance, 4),
        "signal_strength": float(signal["signal_strength"]),
        "strength_band": str(signal["strength_band"]),
        "strength_adx_component": float(signal.get("strength_adx_component", 0.0)),
        "strength_volume_component": float(signal.get("strength_volume_component", 0.0)),
        "strength_ema_component": float(signal.get("strength_ema_component", 0.0)),
        "strength_vwap_component": float(signal.get("strength_vwap_component", 0.0)),
        "strength_trigger_component": float(signal.get("strength_trigger_component", 0.0)),
        "opening_balance": round(balance, 2),
        "paper_status": "OPEN",
    }
    return position, (
        f"Paper ENTRY {position['side']} {symbol} qty={qty} entry={entry:.2f} "
        f"stop={stop:.2f} target={target:.2f} strength={position['signal_strength']:.1f}"
    )


def calculate_charges(entry: float, exit_price: float, qty: int, direction: int, config: PaperConfig) -> dict[str, float]:
    if direction == 1:
        buy_value = entry * qty
        sell_value = exit_price * qty
    else:
        sell_value = entry * qty
        buy_value = exit_price * qty
    turnover = buy_value + sell_value
    brokerage = min(config.brokerage_rate * buy_value, config.brokerage_cap_per_order) + min(
        config.brokerage_rate * sell_value,
        config.brokerage_cap_per_order,
    )
    stt = sell_value * config.stt_sell_side_rate
    exchange = turnover * config.exchange_txn_rate
    sebi = turnover * config.sebi_turnover_rate
    stamp = buy_value * config.stamp_buy_side_rate
    gst = (brokerage + exchange + sebi) * config.gst_rate
    total = brokerage + stt + exchange + sebi + stamp + gst
    return {
        "buy_value": buy_value,
        "sell_value": sell_value,
        "turnover": turnover,
        "brokerage": brokerage,
        "stt": stt,
        "exchange": exchange,
        "sebi": sebi,
        "stamp": stamp,
        "gst": gst,
        "total_charges": total,
    }


def check_exit(position: dict[str, Any], df_1m: pd.DataFrame, config: PaperConfig, current: datetime) -> tuple[dict[str, Any] | None, str]:
    entry_time = pd.Timestamp(position["entry_time"]).to_pydatetime()
    after_entry = df_1m[df_1m["Datetime"] >= entry_time].copy()
    if after_entry.empty:
        return None, "Position open; waiting for candles after entry."

    direction = int(position["direction"])
    stop = float(position["stop"])
    target = float(position["target"])
    entry = float(position["entry"])
    qty = int(position["qty"])
    exit_price = float(after_entry.iloc[-1]["Close"])
    exit_time = pd.Timestamp(after_entry.iloc[-1]["Datetime"]).to_pydatetime()
    exit_reason = None

    for row in after_entry.itertuples(index=False):
        if direction == 1:
            stop_hit = float(row.Low) <= stop
            target_hit = float(row.High) >= target
        else:
            stop_hit = float(row.High) >= stop
            target_hit = float(row.Low) <= target
        if stop_hit and target_hit:
            exit_price = stop
            exit_time = pd.Timestamp(row.Datetime).to_pydatetime()
            exit_reason = "stop_same_minute"
            break
        if stop_hit:
            exit_price = stop
            exit_time = pd.Timestamp(row.Datetime).to_pydatetime()
            exit_reason = "stop"
            break
        if target_hit:
            exit_price = target
            exit_time = pd.Timestamp(row.Datetime).to_pydatetime()
            exit_reason = "target"
            break

    if exit_reason is None and current.time() >= parse_time(EOD_EXIT_TIME):
        exit_price = float(after_entry.iloc[-1]["Close"])
        exit_time = pd.Timestamp(after_entry.iloc[-1]["Datetime"]).to_pydatetime()
        exit_reason = "eod"

    if exit_reason is None:
        return None, "Position open; stop/target not hit."

    exit_price = apply_slippage(exit_price, direction, "exit", config)
    gross_pnl = direction * (exit_price - entry) * qty
    charges = calculate_charges(entry, exit_price, qty, direction, config)
    net_pnl = gross_pnl - charges["total_charges"]
    opening_balance = float(position["opening_balance"])
    closing_balance = opening_balance + net_pnl
    trade = {
        **position,
        "exit_time": exit_time.isoformat(),
        "exit": round(exit_price, 4),
        "exit_reason": exit_reason,
        "gross_pnl": round(gross_pnl, 2),
        **{key: round(value, 2) for key, value in charges.items()},
        "net_pnl": round(net_pnl, 2),
        "net_return_pct": round((net_pnl / opening_balance) * 100, 4) if opening_balance else 0.0,
        "closing_balance": round(closing_balance, 2),
        "paper_status": "CLOSED",
    }
    return trade, (
        f"Paper EXIT {position['symbol']} {exit_reason} exit={exit_price:.2f} "
        f"net_pnl={net_pnl:.2f} balance={closing_balance:.2f}"
    )


def is_data_fresh(df_1m: pd.DataFrame, current: datetime, config: PaperConfig) -> tuple[bool, str]:
    if df_1m.empty:
        return False, "No 1MIN data available."
    latest = pd.Timestamp(df_1m["Datetime"].max()).to_pydatetime()
    expected = latest_expected_closed_minute(current)
    if expected is None:
        return True, "Outside active market freshness window."
    stale_after = expected - timedelta(minutes=max(0, config.max_data_staleness_minutes))
    if latest < stale_after:
        return False, f"Stale 1MIN data: latest={latest}, expected around={expected}."
    return True, "Data fresh enough."


def run_once(
    symbols: Iterable[str] = DEFAULT_SYMBOLS,
    *,
    data_folder: str | Path = DEFAULT_DATA_FOLDER,
    report_folder: str | Path = DEFAULT_REPORT_DIR,
    config: PaperConfig | None = None,
    reset: bool = False,
    current: datetime | None = None,
) -> list[str]:
    config = config or PaperConfig()
    current = current or now_ist()
    cleaned_symbols = [clean_symbol(symbol) for symbol in symbols]
    state = load_state(cleaned_symbols, config, report_folder, reset=reset)
    messages: list[str] = []

    for symbol in cleaned_symbols:
        df_1m = load_candles(symbol, Main.TIMEFRAME_1MIN, data_folder)
        df_5m = load_candles(symbol, Main.TIMEFRAME_5MIN, data_folder)
        fresh, freshness_message = is_data_fresh(df_1m, current, config)
        if not fresh:
            append_event("DATA_STALE", symbol, freshness_message, report_folder)
            messages.append(f"{symbol}: {freshness_message}")
            continue
        if df_5m.empty:
            msg = f"{symbol}: No 5MIN data available for strategy evaluation."
            append_event("NO_DATA", symbol, msg, report_folder)
            messages.append(msg)
            continue

        open_position_map = state.setdefault("open_positions", {})
        if symbol in open_position_map:
            trade, exit_msg = check_exit(open_position_map[symbol], df_1m, config, current)
            if trade is None:
                append_event("HOLD", symbol, exit_msg, report_folder)
                messages.append(f"{symbol}: {exit_msg}")
                continue
            append_csv(trades_path(report_folder), trade)
            append_event("EXIT", symbol, exit_msg, report_folder, trade)
            state["balance"] = float(trade["closing_balance"])
            del open_position_map[symbol]
            messages.append(f"{symbol}: {exit_msg}")
            continue

        if len(open_position_map) >= config.max_open_positions:
            msg = f"Portfolio already has {len(open_position_map)} open position(s); skipping new {symbol} entries."
            append_event("CAPACITY_SKIP", symbol, msg, report_folder)
            messages.append(f"{symbol}: {msg}")
            continue

        signal = candidate_signal_for_today(df_5m, current)
        if signal is None:
            msg = "No gold strategy signal yet."
            append_event("NO_SIGNAL", symbol, msg, report_folder)
            messages.append(f"{symbol}: {msg}")
            continue

        key = signal_key(symbol, signal)
        if key in state.get("seen_signal_keys", []):
            msg = f"Signal already processed: {key}"
            append_event("DUPLICATE_SIGNAL", symbol, msg, report_folder)
            messages.append(f"{symbol}: {msg}")
            continue

        entry_time = pd.Timestamp(signal["entry_time"]).to_pydatetime()
        entry_candle = latest_entry_candle(df_1m, entry_time, config.entry_grace_minutes)
        if entry_candle is None:
            if current > entry_time + timedelta(minutes=config.entry_grace_minutes):
                state.setdefault("missed_signal_keys", []).append(key)
                state.setdefault("seen_signal_keys", []).append(key)
                msg = f"Missed signal entry window for {entry_time}."
                append_event("MISSED_SIGNAL", symbol, msg, report_folder, signal.to_dict())
                messages.append(f"{symbol}: {msg}")
            else:
                msg = signal_alert_message(symbol, signal, "SIGNAL HIT; waiting for entry candle")
                append_event("PENDING_ENTRY", symbol, msg, report_folder, signal.to_dict())
                messages.append(msg)
            continue

        position, entry_msg = open_position(symbol, signal, entry_candle, state, config)
        state.setdefault("seen_signal_keys", []).append(key)
        if position is None:
            append_event("ENTRY_REJECTED", symbol, entry_msg, report_folder, signal.to_dict())
            messages.append(f"{symbol}: {entry_msg}")
            continue
        open_position_map[symbol] = position
        append_event("ENTRY", symbol, entry_msg, report_folder, position)
        messages.append(f"{symbol}: {entry_msg}")

    state["last_status"] = " | ".join(messages)
    save_state(state, report_folder)
    write_report(report_folder, config, current)
    return messages


def write_report(
    report_folder: str | Path = DEFAULT_REPORT_DIR,
    config: PaperConfig | None = None,
    current: datetime | None = None,
) -> Path:
    config = config or PaperConfig()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write paper trading .xlsx reports.") from exc

    report_folder = report_dir(report_folder)
    trades = read_report_csv(trades_path(report_folder))
    events = read_report_csv(events_path(report_folder))
    state = json.loads(state_path(report_folder).read_text(encoding="utf-8")) if state_path(report_folder).exists() else {}
    open_positions = list((state.get("open_positions") or {}).values())

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_trades = wb.create_sheet("Trade Ledger")
    ws_events = wb.create_sheet("Events")
    ws_positions = wb.create_sheet("Open Positions")
    ws_assumptions = wb.create_sheet("Assumptions")

    header_fill = PatternFill("solid", fgColor="1F2937")
    subheader_fill = PatternFill("solid", fgColor="E5E7EB")
    white = "FFFFFF"
    green = "166534"
    red = "991B1B"

    def style_title(ws, title: str, end_col: int = 6) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        cell = ws.cell(1, 1, title)
        cell.fill = header_fill
        cell.font = Font(color=white, bold=True, size=14)
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 24

    def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
        if df.empty:
            ws.cell(start_row, start_col, "No rows yet")
            return
        for col_idx, col_name in enumerate(df.columns, start_col):
            cell = ws.cell(start_row, col_idx, col_name)
            cell.fill = header_fill
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(wrap_text=True)
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row, start_col):
                ws.cell(row_idx, col_idx, _excel_value(value))
        ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate

    style_title(ws_summary, "Gold Strategy Live Paper Trading")
    summary_rows = summary_rows_from_state(state, trades, config)
    ws_summary.append(["Metric", "Value"])
    for row in summary_rows:
        ws_summary.append(row)
    for cell in ws_summary[2]:
        cell.fill = subheader_fill
        cell.font = Font(bold=True)
    for row in ws_summary.iter_rows(min_row=3, max_col=2):
        if row[0].value in ("Net P/L", "Open Unrealized P/L") and isinstance(row[1].value, (int, float)):
            row[1].font = Font(color=green if row[1].value >= 0 else red, bold=True)

    write_dataframe(ws_trades, trades)
    write_dataframe(ws_events, events.tail(250).reset_index(drop=True) if not events.empty else events)
    write_dataframe(ws_positions, pd.DataFrame(open_positions))

    style_title(ws_assumptions, "Paper Trading Assumptions", end_col=4)
    assumptions = [
        ["Initial Balance", config.initial_balance, "Rs", "Portfolio starting balance"],
        ["Leverage", config.leverage, "x", "Whole-share quantity uses balance * leverage / price"],
        ["Max Open Positions", config.max_open_positions, "", "Default is conservative single-position mode"],
        ["Entry Grace", config.entry_grace_minutes, "minutes", "Miss signal if matching entry candle is too old"],
        ["Max Data Staleness", config.max_data_staleness_minutes, "minutes", "No trades if LiveTick CSV is stale"],
        ["Brokerage Rate", config.brokerage_rate, "%", "Per order, capped"],
        ["Brokerage Cap", config.brokerage_cap_per_order, "Rs", "Applied separately on buy and sell"],
        ["STT Sell Side", config.stt_sell_side_rate, "%", "Equity intraday STT"],
        ["Exchange Txn", config.exchange_txn_rate, "%", "NSE transaction charge"],
        ["SEBI Turnover", config.sebi_turnover_rate, "%", "SEBI turnover fee"],
        ["Stamp Buy Side", config.stamp_buy_side_rate, "%", "Equity intraday stamp duty"],
        ["GST", config.gst_rate, "%", "On brokerage + exchange + SEBI"],
        ["Slippage", config.slippage_bps, "bps", "Optional simulation slippage"],
    ]
    ws_assumptions.append([])
    ws_assumptions.append(["Assumption", "Value", "Unit", "Note"])
    for row in assumptions:
        ws_assumptions.append(row)
    for cell in ws_assumptions[3]:
        cell.fill = subheader_fill
        cell.font = Font(bold=True)

    for ws in [ws_summary, ws_trades, ws_events, ws_positions, ws_assumptions]:
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws[col_letter]:
                max_len = max(max_len, min(45, len(str(cell.value)) if cell.value is not None else 0))
            ws.column_dimensions[col_letter].width = max_len + 2
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    path = report_xlsx_path(report_folder, current)
    return _save_workbook_safely(wb, path)


def _save_workbook_safely(wb, path: str | Path) -> Path:
    primary = Path(path)
    try:
        wb.save(primary)
        return primary
    except PermissionError:
        fallback = fallback_report_xlsx_path(primary)
        try:
            wb.save(fallback)
            print(
                f"Report file is locked, wrote live report copy instead: {fallback}",
                flush=True,
            )
            return fallback
        except PermissionError:
            stamped = primary.with_name(f"{primary.stem}_{now_ist().strftime('%H%M%S')}{primary.suffix}")
            wb.save(stamped)
            print(
                f"Report files are locked, wrote timestamped report copy instead: {stamped}",
                flush=True,
            )
            return stamped


def read_report_csv(path: str | Path) -> pd.DataFrame:
    frames = []
    for candidate in (Path(path), fallback_report_file_path(path)):
        if candidate.exists() and candidate.stat().st_size > 0:
            frames.append(pd.read_csv(candidate))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)


def summary_rows_from_state(state: dict[str, Any], trades: pd.DataFrame, config: PaperConfig) -> list[list[Any]]:
    balance = float(state.get("balance", config.initial_balance))
    open_positions = state.get("open_positions") or {}
    if trades.empty:
        closed = 0
        wins = 0
        net_pnl = 0.0
        total_charges = 0.0
        max_drawdown = 0.0
    else:
        closed = int(len(trades))
        wins = int((pd.to_numeric(trades.get("net_pnl", 0), errors="coerce") > 0).sum())
        net_pnl = float(pd.to_numeric(trades.get("net_pnl", 0), errors="coerce").fillna(0).sum())
        total_charges = float(pd.to_numeric(trades.get("total_charges", 0), errors="coerce").fillna(0).sum())
        balances = pd.to_numeric(trades.get("closing_balance", pd.Series(dtype=float)), errors="coerce").dropna()
        if balances.empty:
            max_drawdown = 0.0
        else:
            max_drawdown = float((balances / balances.cummax() - 1).min() * 100)

    return [
        ["Session Date", state.get("session_date", "")],
        ["Started At", state.get("started_at", "")],
        ["Symbols", ", ".join(state.get("symbols", []))],
        ["Initial Balance", config.initial_balance],
        ["Current Balance", round(balance, 2)],
        ["Net P/L", round(net_pnl, 2)],
        ["Closed Trades", closed],
        ["Win Rate", round((wins / closed) * 100, 2) if closed else 0.0],
        ["Total Charges", round(total_charges, 2)],
        ["Max Drawdown %", round(max_drawdown, 2)],
        ["Open Positions", len(open_positions)],
        ["Last Status", state.get("last_status", "")],
    ]


def _excel_value(value: Any) -> Any:
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if pd.isna(value):
        return None
    return value


def is_signal_console_message(message: str) -> bool:
    return "SIGNAL HIT" in message or "Paper ENTRY" in message


def colorize_green(text: str) -> str:
    return f"{ANSI_GREEN}{text}{ANSI_RESET}"


def print_status_line(current: datetime, message: str) -> None:
    line = f"{current.isoformat(timespec='seconds')} {message}"
    if is_signal_console_message(message):
        line = colorize_green(line)
    print(line, flush=True)


def live_tick_session_running() -> tuple[bool, dict[str, Any] | None]:
    session = load_session("live_tick")
    if not session:
        return False, None
    return is_pid_running(session.get("pid")), session


def live_tick_session_symbols(session: dict[str, Any] | None) -> set[str]:
    metadata = (session or {}).get("metadata") or {}
    symbols = metadata.get("symbols") or []
    return {clean_symbol(symbol) for symbol in symbols}


def _run_managed_live_tick(
    symbols: list[str],
    data_folder: Path,
    tick_root: Path,
    errors: list[BaseException],
) -> None:
    try:
        from LiveTick import LiveTick

        LiveTick(symbols, output_folder=data_folder, tick_root=tick_root)
    except BaseException as exc:
        errors.append(exc)


def ensure_live_tick_feed(
    symbols: Iterable[str],
    *,
    data_folder: str | Path = DEFAULT_DATA_FOLDER,
    tick_root: str | Path = DEFAULT_TICK_ROOT,
    report_folder: str | Path = DEFAULT_REPORT_DIR,
    manage_live_tick: bool = True,
) -> tuple[bool, threading.Thread | None, list[BaseException], list[str]]:
    cleaned_symbols = [clean_symbol(symbol) for symbol in symbols]
    running, session = live_tick_session_running()
    if running:
        session_symbols = live_tick_session_symbols(session)
        missing = sorted(set(cleaned_symbols) - session_symbols) if session_symbols else []
        if missing:
            message = (
                f"Existing LiveTick PID {session.get('pid')} is running, but it does not list "
                f"{', '.join(missing)} in its session metadata. Reusing it; missing symbols may stay stale."
            )
            append_event("LIVE_FEED_SYMBOL_MISMATCH", "PORTFOLIO", message, report_folder)
            return False, None, [], [message]

        message = f"Reusing LiveTick session PID {session.get('pid')} for {', '.join(cleaned_symbols)}."
        append_event("LIVE_FEED_REUSE", "PORTFOLIO", message, report_folder)
        return False, None, [], [message]

    if not manage_live_tick:
        message = "LiveTick auto-start is disabled; paper trader will only use existing local CSV candles."
        append_event("LIVE_FEED_DISABLED", "PORTFOLIO", message, report_folder)
        return False, None, [], [message]

    errors: list[BaseException] = []
    thread = threading.Thread(
        target=_run_managed_live_tick,
        args=([*cleaned_symbols], Path(data_folder), Path(tick_root), errors),
        name=f"GoldPaperLiveTick-{','.join(cleaned_symbols)}",
        daemon=True,
    )
    thread.start()
    message = f"Started managed LiveTick feed for {', '.join(cleaned_symbols)}."
    append_event("LIVE_FEED_START", "PORTFOLIO", message, report_folder)
    return True, thread, errors, [message]


def wait_for_managed_live_tick(
    thread: threading.Thread | None,
    errors: list[BaseException],
    *,
    timeout_seconds: int = 20,
) -> str:
    if thread is None:
        return "Using existing live feed state."

    deadline = time.monotonic() + max(1, int(timeout_seconds))
    while time.monotonic() < deadline:
        if errors:
            return f"LiveTick startup failed: {errors[-1]}"
        running, session = live_tick_session_running()
        if running:
            return f"LiveTick session is running with PID {session.get('pid')}."
        if not thread.is_alive():
            return "LiveTick completed startup reconciliation without opening an active stream."
        time.sleep(1)
    return "LiveTick is still starting; paper loop will wait for fresh CSV candles."


def interruptible_sleep(
    seconds: int,
    *,
    stop_requested: Callable[[], bool],
    live_errors: list[BaseException],
    reported_live_errors: Callable[[], int],
) -> None:
    deadline = time.monotonic() + max(0, int(seconds))
    while time.monotonic() < deadline and not stop_requested():
        if len(live_errors) > reported_live_errors():
            return
        time.sleep(min(1, deadline - time.monotonic()))


def GoldPaperTrade(
    symbols: Iterable[str] = DEFAULT_SYMBOLS,
    *,
    data_folder: str | Path = DEFAULT_DATA_FOLDER,
    report_folder: str | Path = DEFAULT_REPORT_DIR,
    tick_root: str | Path = DEFAULT_TICK_ROOT,
    poll_seconds: int = 30,
    duration_minutes: int = 240,
    reset: bool = False,
    manage_live_tick: bool = True,
    config: PaperConfig | None = None,
) -> None:
    config = config or PaperConfig()
    symbols = [clean_symbol(symbol) for symbol in symbols]
    stop_requested = False
    live_owned = False
    live_thread: threading.Thread | None = None
    live_errors: list[BaseException] = []
    reported_live_errors = 0

    def stop_loop() -> None:
        nonlocal stop_requested
        stop_requested = True

    metadata = {
        "symbols": symbols,
        "report_folder": str(report_folder),
        "paper_config": asdict(config),
        "manage_live_tick": manage_live_tick,
    }
    try:
        with RuntimeSession("paper_trader", metadata=metadata, on_stop_requested=stop_loop):
            start_time = now_ist()
            single_check_only = should_run_single_offmarket_check(start_time)
            if should_start_live_tick(start_time, manage_live_tick):
                live_owned, live_thread, live_errors, startup_messages = ensure_live_tick_feed(
                    symbols,
                    data_folder=data_folder,
                    tick_root=tick_root,
                    report_folder=report_folder,
                    manage_live_tick=manage_live_tick,
                )
            else:
                message = offmarket_live_tick_message(start_time)
                append_event("LIVE_FEED_SKIPPED", "PORTFOLIO", message, report_folder)
                startup_messages = [message]
            for message in startup_messages:
                print(message, flush=True)
            if live_thread is not None:
                readiness = wait_for_managed_live_tick(live_thread, live_errors)
                append_event("LIVE_FEED_READY", "PORTFOLIO", readiness, report_folder)
                print(readiness, flush=True)

            end_at = now_ist() + timedelta(minutes=max(1, int(duration_minutes)))
            first = True
            try:
                while now_ist() <= end_at and not stop_requested:
                    current = now_ist()
                    if reported_live_errors < len(live_errors):
                        for exc in live_errors[reported_live_errors:]:
                            append_event("LIVE_FEED_ERROR", "PORTFOLIO", str(exc), report_folder)
                            print(f"{current.isoformat(timespec='seconds')} LIVE_FEED_ERROR: {exc}", flush=True)
                        reported_live_errors = len(live_errors)
                        if is_market_session(current):
                            print("Stopping paper trader because managed LiveTick failed during market hours.", flush=True)
                            break

                    try:
                        messages = run_once(
                            symbols,
                            data_folder=data_folder,
                            report_folder=report_folder,
                            config=config,
                            reset=reset and first,
                            current=current,
                        )
                        for message in messages:
                            print_status_line(current, message)
                        if single_check_only:
                            break
                    except Exception as exc:
                        for symbol in symbols:
                            append_event("ERROR", symbol, str(exc), report_folder)
                        write_report(report_folder, config, current)
                        print_status_line(current, f"ERROR: {exc}")
                        if single_check_only:
                            break
                    first = False
                    interruptible_sleep(
                        max(5, int(poll_seconds)),
                        stop_requested=lambda: stop_requested,
                        live_errors=live_errors,
                        reported_live_errors=lambda: reported_live_errors,
                    )
            except KeyboardInterrupt:
                stop_requested = True
                print("Keyboard interrupt received; shutting down paper trader gracefully.", flush=True)
    finally:
        if live_owned and live_thread is not None and live_thread.is_alive():
            request_stop("live_tick", reason="paper_trader_shutdown")
            try:
                live_thread.join(timeout=10)
            except KeyboardInterrupt:
                print("LiveTick stop was requested; exiting without waiting for the socket thread.", flush=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run live paper trading for the gold intraday strategy.")
    parser.add_argument("--symbols", nargs="+", default=list(DEFAULT_SYMBOLS))
    parser.add_argument("--data-folder", default=str(DEFAULT_DATA_FOLDER))
    parser.add_argument("--report-folder", default=str(DEFAULT_REPORT_DIR))
    parser.add_argument("--tick-root", default=str(DEFAULT_TICK_ROOT))
    parser.add_argument("--poll-seconds", type=int, default=30)
    parser.add_argument("--duration-minutes", type=int, default=240)
    parser.add_argument("--initial-balance", type=float, default=1000.0)
    parser.add_argument("--leverage", type=float, default=5.0)
    parser.add_argument("--reset", action="store_true")
    parser.add_argument("--once", action="store_true")
    parser.add_argument("--no-manage-live-tick", action="store_true")
    parser.add_argument("--stop", action="store_true")
    parser.add_argument("--status", action="store_true")
    args = parser.parse_args()

    if args.stop:
        path = request_stop("paper_trader")
        print(f"Stop requested: {path}")
        return
    if args.status:
        print(json.dumps(load_session("paper_trader") or {}, indent=2, default=str))
        return

    config = PaperConfig(initial_balance=args.initial_balance, leverage=args.leverage)
    if args.once:
        for message in run_once(
            args.symbols,
            data_folder=args.data_folder,
            report_folder=args.report_folder,
            config=config,
            reset=args.reset,
        ):
            print(message)
    else:
        GoldPaperTrade(
            args.symbols,
            data_folder=args.data_folder,
            report_folder=args.report_folder,
            tick_root=args.tick_root,
            poll_seconds=args.poll_seconds,
            duration_minutes=args.duration_minutes,
            reset=args.reset,
            manage_live_tick=not args.no_manage_live_tick,
            config=config,
        )


if __name__ == "__main__":
    main()
